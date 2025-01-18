import os
import pickle
import pandas as pd
import openpyxl as pxl
from openpyxl.utils import get_column_letter


def save_list_to_txt(my_list, file_path):
    # os.path.join(dir, file_path)
    try:
        with open(file_path, 'w') as txt_file:
            for item in my_list:
                txt_file.write(str(item) + '\n')
        print(f"List saved to {file_path} successfully.")
    except Exception as e:
        print(f"Error: {e}")

def load_list_from_txt(file_path):
    # os.path.join(dir, file_path)
    try:
        with open(file_path, 'r') as txt_file:
            my_list = txt_file.read().splitlines()
        print(f"List loaded from {file_path} successfully.")
        return my_list
    except Exception as e:
        print(f"Error: {e}")
        return None


def load_dict_pickles_and_concatenate(directory) -> dict: 
    # Initialize an empty dictionary to store the concatenated data.
    concatenated_dict = {}

    # List all files in the directory.
    files = os.listdir(directory)

    # Loop through each file in the directory.
    for filename in files:
        # Check if the file is a pickle file (ends with .pkl).
        if filename.endswith(".pickle"):
            file_path = os.path.join(directory, filename)
            
            # Load the pickle file and merge it into the concatenated_dict.
            with open(file_path, 'rb') as file:
                data = pickle.load(file)

                if isinstance(data, dict):
                    concatenated_dict.update(data)
                # else:
                    # concatenated_dict = pd.concat([concatenated_dict, data], axis=0)

    return concatenated_dict

def load_df_pickles_and_concatenate(directory) -> pd.DataFrame: 
    # Initialize an empty dictionary to store the concatenated data.
    df = pd.DataFrame([])

    # List all files in the directory.
    files = os.listdir(directory)

    # Loop through each file in the directory.
    for filename in files:
        # Check if the file is a pickle file (ends with .pkl).
        if filename.endswith(".pickle"):
            file_path = os.path.join(directory, filename)
            
            # Load the pickle file and merge it into the concatenated_dict.
            with open(file_path, 'rb') as file:
                data = pickle.load(file)
                df = pd.concat([df, data], axis=0)

    return df



# Save the input dataframe to the specified sheet name of filename file
def saveAsNewSheetToExistingFile(filename,newDF,newSheetName):
    if os.path.exists(filename):

        excel_book = pxl.load_workbook(filename)
        
#         print(excel_book.sheetnames)
#         if 'cp-cd' in excel_book.sheetnames:
#             print('ghalate')

        if newSheetName in excel_book.sheetnames:
            del excel_book[newSheetName]
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Your loaded workbook is set as the "base of work"
            writer.book = excel_book

            # Loop through the existing worksheets in the workbook and map each title to\
            # the corresponding worksheet (that is, a dictionary where the keys are the\
            # existing worksheets' names and the values are the actual worksheets)
#             print(excel_book.worksheets)
            writer.sheets = {worksheet.title: worksheet for worksheet in excel_book.worksheets if newSheetName not in worksheet}

            # Write the new data to the file without overwriting what already exists
            newDF.to_excel(writer, newSheetName)

            # Save the file
            writer.save()
    else:
        newDF.to_excel(filename, newSheetName)
        
    print(newSheetName,' saved!')
    return




def write_dataframe_to_excel(filename, sheet_name, dataframe,append_new_data_if_sheet_exists=True):

    if os.path.exists(filename):
        print(f'file {filename} exists')
        try:
            # Load the existing Excel file or create a new one
            try:
                wb = pxl.load_workbook(filename)
                sheet_exists = sheet_name in wb.sheetnames
            except FileNotFoundError:
                wb = pxl.load_workbook(filename)
                sheet_exists = False

            # Check if the sheet_name already exists
            if sheet_exists:
                
                if append_new_data_if_sheet_exists==False:
                    print(f'Sheet "{sheet_name}" already exists in the Excel file. Not adding new data to the existing sheet.')
                    return

                print(f'Sheet "{sheet_name}" already exists in the Excel file. Adding new data to the existing sheet.')
                ws = wb[sheet_name]
                startrow = ws.max_row + 1
            else:
                ws = wb.create_sheet(sheet_name)
                startrow = 1

            # Convert DataFrame to a list of lists to write to Excel
            data_to_write = dataframe.values.tolist()

                                # Write the column headers to the sheet
            if not sheet_exists:
                for col_idx, column in enumerate(dataframe.columns, 1):
                    col_letter = get_column_letter(col_idx)
                    ws[f"{col_letter}{startrow}"] = column
                startrow += 1

            # # Write the data to the specified sheet_name
            for row_data in data_to_write:
                ws.append(row_data)
            # Save the changes
            wb.save(filename)

            print(f'DataFrame successfully written to sheet "{sheet_name}" in Excel file "{filename}".')
        except Exception as e:
            print(f'Error: {e}')

    else:
        print(f'Excel file "{filename}" does not exist. Creating new file with sheet "{sheet_name}".')
        dataframe.to_excel(filename, sheet_name,index=False)
        
    print(sheet_name,' saved!')
    return