import sys
from datetime import datetime
import os
import torch
import numpy as np
import random
import logging
import argparse
from data_layer.data_utils import get_cp_path, get_cp_dir
from utils.configuration import DataArguments, GeneralArguments, ModelArguments, EvalArguments, MoaArguments
from utils.global_variables import DS_INFO_DICT
from utils.global_variables import ABRVS
import transformers
import pandas as pd
import openpyxl as pxl
from openpyxl.utils import get_column_letter
import os
import json

from argparse import ArgumentParser
from pytorch_lightning import seed_everything
import pickle


################################################################################

def revise_exp_name(configs):
    if configs.model.tune_hyperparams:
        configs.general.exp_name += f'_{ABRVS["tune"]}'
    if configs.data.norm_method == 'mad_robustize':
        configs.general.exp_name += f'_{ABRVS[configs.data.norm_method]}'
    # if configs.model.l2_lambda>0.01:
        # configs.general.exp_name += f'_l2_{configs.model.l2_lambda}'
    if configs.model.deep_decoder:
        configs.general.exp_name += f'_dd'
        
    if configs.model.encoder_type == 'shallow':
        configs.general.exp_name += f'_se'
    elif configs.model.encoder_type == 'deep':
        configs.general.exp_name += f'_de'
    return configs.general.exp_name 

################################################################################

def add_exp_suffix(profile_type, by_dose,normalize_by_all=False, z_trim=None, min_max_norm=False):

    suffix = ''
    
    if profile_type == 'augmented':
        suffix+= '_a'
    elif profile_type == 'normalized_variable_selected':
        suffix+= '_nvs'
    if by_dose:
        suffix+= '_d'
    if normalize_by_all:
        suffix+= '_ba'
    if z_trim is not None:
        suffix+= f'_z{str(z_trim)}'
    if min_max_norm:
        suffix+= '_mm' 
        
    return suffix


################################################################################

def set_configs(exp_name = ''):
    
    parser = transformers.HfArgumentParser((GeneralArguments, DataArguments,ModelArguments, EvalArguments,MoaArguments))
    general_args, data_args,model_args, eval_args, moa_args = parser.parse_args_into_dataclasses()
    
    configs = argparse.Namespace(**{
        'general': general_args,
        'model': model_args,
        'data': data_args,
        'eval':eval_args,
        'moa':moa_args
    })
    if len(exp_name)>0:
        configs.general.exp_name = exp_name
    exp_name = revise_exp_name(configs)
    configs.general.exp_name = exp_name
    configs.general.output_exp_dir = f"{configs.general.base_dir}/anomaly_output/{configs.general.dataset}/{configs.data.modality}/{configs.general.exp_name}/"
    # configs = set_paths(configs)

    # configs_from_file = get_configs(configs.general.output_exp_dir)
    # if configs_from_file is not None:
    #     configs = configs_from_file
    set_seed(configs.general.seed)
        # return None

    return configs

################################################################################

def save_configs(configs):
    config_path = os.path.join(configs.general.output_exp_dir,'args.pkl')
    
    with open(config_path, 'wb') as f:
        pickle.dump(configs, f)
    
    # save text file with all params    
    with open(os.path.join(configs.general.output_exp_dir, 'params.txt'), 'w') as f:
        # skip line for each parameter
        for k,v in configs.__dict__.items():
            f.write(f'{k}:\n')

            for k2,v2 in v.__dict__.items():
                # add indetation for each sub-parameter
                f.write(f'     {k2}: {v2}\n')


################################################################################

def get_configs(exp_dir):
    # path = get_cp_dir(DATASET_ROOT_DIR,DS_INFO_DICT[DATASET]['name'],exp_name)
    config_path = os.path.join(exp_dir,'args.pkl')
    
    if os.path.exists(config_path):  
        parser = ArgumentParser()

        args, unknown = parser.parse_known_args()
        # args = parser.parse_args()  

        with open(config_path, 'rb') as f:
            return pickle.load(f)
        
    else:
        return None
    

################################################################################

def set_configs_jupyter(exp_name = ''):

    parser = transformers.HfArgumentParser((GeneralArguments, DataArguments,ModelArguments, EvalArguments,MoaArguments))
    general_args = GeneralArguments()
    data_args = DataArguments()
    model_args = ModelArguments()
    eval_args = EvalArguments()
    moa_args = MoaArguments()

    # general_args, data_args,model_args, eval_args = parser.parse_args_into_dataclasses()
    # general_args, data_args,model_args, eval_args = parser.parse_known_args()
    # configs = set_paths(configs)
    # assert_configs(configs)
    # configs.general.exp_name = exp_name


    # cpds_med_scores, null_distribution_medians = calc_reproducibility(configs)
    # general_args.flow
    configs = argparse.Namespace(**{
            'general': general_args,
            'model': model_args,
            'data': data_args,
            'eval':eval_args,
            'moa':moa_args
        })
    
    if len(exp_name)>0:
        configs.general.exp_name = exp_name
    exp_name = revise_exp_name(configs)
    configs.general.exp_name = exp_name
    # configs = set_paths(configs)

    # configs_from_file = get_configs(configs.general.output_exp_dir)
    # if configs_from_file is not None:
        # configs = configs_from_file

    set_seed(configs.general.seed)
    
    return configs

def set_logger(configs):
    if configs.general.debug_mode == True:
        logging.basicConfig(filename=os.path.join(configs.general.res_dir, 'debug_log.log'),
                            datefmt='%H:%M:%S',
                            level=logging.INFO)
    else:
        log_dir = os.path.join(configs.general.res_dir,'logs', configs.general.flow)
        os.makedirs(log_dir, exist_ok=True)
        logging.basicConfig(filename=os.path.join(log_dir, f'log_exp_num={configs.general.exp_num}.log'),
                            filemode='a',
                            format='%(asctime)s,%(msecs)d %(name)s %(levelname)s %(message)s',
                            datefmt='%H:%M:%S',
                            level=logging.INFO)
        logging.getLogger().addHandler(logging.StreamHandler(sys.stdout))
    configs.general.logger = logging.getLogger(__name__)
    return configs


def assert_configs(configs):
    assert configs.general.dataset in ['CDRP-bio','CDRP', 'LINCS', 'LUAD', 'TAORF'], "dataset not supported"
    assert configs.data.modality in ['CellPainting', 'L1000'], "modality not supported"
    assert configs.data.profile_type in ['augmented', 'normalized', 'normalized_variable_selected'], "profile_type not supported"
    # assert configs.general.flow in ['run_ad', 'calc_metrics', 'analyze'], "flow not supported"
    assert configs.general.exp_name != '', "exp_name is empty"
    assert configs.general.base_dir != '', "base_dir is empty"
    # assert configs.general.output_exp_dir != '', "output_dir is empty"
    assert configs.general.res_dir != '', "res_dir is empty"

def set_paths(configs):
    configs.general.data_path = get_cp_path(configs.general.base_dir, configs.general.dataset,configs.data.profile_type)
    configs.general.processed_data_dir = get_cp_dir(configs.general.base_dir, configs.general.dataset, configs.data.profile_type,processed=True)
    configs.general.output_exp_dir = get_cp_dir(configs.general.base_dir, configs.general.dataset, configs.data.profile_type, exp_name=configs.general.exp_name,processed=True)
    configs.general.res_dir = f"{configs.general.base_dir}/results/{configs.general.dataset}/{configs.data.modality}/{configs.general.exp_name}/"
    configs.general.fig_dir = f"{configs.general.res_dir}/figs"
    configs.model.ckpt_dir = f"{configs.general.res_dir}/ckpt"
    configs.model.tb_logs_dir = f"{configs.general.res_dir}/logs"

    os.makedirs(configs.general.output_exp_dir, exist_ok=True)
    os.makedirs(configs.general.res_dir, exist_ok=True)
    os.makedirs(configs.general.fig_dir, exist_ok=True)

    assert_configs(configs)
    if not hasattr(configs.general,'logger'):
        configs = set_logger(configs)
    # os.makedirs(configs.general.processed_data_dir, exist_ok=True)

    return configs
    

def output_path(configs):
    output_path = 'Not created'
    if configs.general.debug_mode == False:
        now = datetime.now()
        output_name = now.strftime("%d_%m_%Y____%H_%M_%S")
        output_path = os.path.join(
            configs.general.output_exp_dir,
            configs.model.name,
            configs.data.dataset_name,
            output_name
        )
        if configs.general.debug_mode == False and os.path.exists(output_path) == False:
            print('path is: {}'.format(output_path))
            os.makedirs(output_path, exist_ok=True)

    return output_path


def set_seed(seed):
    # torch.manual_seed(seed)
    # np.random.seed(seed)
    # random.seed(seed)
    seed_everything(seed)


# Save the input dataframe to the specified sheet name of filename file
def saveAsNewSheetToExistingFile(filename,newDF,newSheetName):
    if os.path.exists(filename):

        excel_book = pxl.load_workbook(filename)
        
#         print(excel_book.sheetnames)
#         if 'cp-cd' in excel_book.sheetnames:
#             print('ghalate')

        if newSheetName in excel_book.sheetnames:
            del excel_book[newSheetName]
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Your loaded workbook is set as the "base of work"
            writer.book = excel_book

            # Loop through the existing worksheets in the workbook and map each title to\
            # the corresponding worksheet (that is, a dictionary where the keys are the\
            # existing worksheets' names and the values are the actual worksheets)
#             print(excel_book.worksheets)
            writer.sheets = {worksheet.title: worksheet for worksheet in excel_book.worksheets if newSheetName not in worksheet}

            # Write the new data to the file without overwriting what already exists
            newDF.to_excel(writer, newSheetName)

            # Save the file
            writer.save()
    else:
        newDF.to_excel(filename, newSheetName)
        
    print(newSheetName,' saved!')
    return




def write_dataframe_to_excel(filename, sheet_name, dataframe,append_new_data_if_sheet_exists=True):

    if os.path.exists(filename):
        print(f'file {filename} exists')
        try:
            # Load the existing Excel file or create a new one
            try:
                wb = pxl.load_workbook(filename)
                sheet_exists = sheet_name in wb.sheetnames
            except FileNotFoundError:
                wb = pxl.load_workbook(filename)
                sheet_exists = False

            # Check if the sheet_name already exists
            if sheet_exists:
                
                if append_new_data_if_sheet_exists==False:
                    print(f'Sheet "{sheet_name}" already exists in the Excel file. Not adding new data to the existing sheet.')
                    return

                print(f'Sheet "{sheet_name}" already exists in the Excel file. Adding new data to the existing sheet.')
                ws = wb[sheet_name]
                startrow = ws.max_row + 1
            else:
                ws = wb.create_sheet(sheet_name)
                startrow = 1

            # Convert DataFrame to a list of lists to write to Excel
            data_to_write = dataframe.values.tolist()

                                # Write the column headers to the sheet
            if not sheet_exists:
                for col_idx, column in enumerate(dataframe.columns, 1):
                    col_letter = get_column_letter(col_idx)
                    ws[f"{col_letter}{startrow}"] = column
                startrow += 1

            # # Write the data to the specified sheet_name
            # for row_data in data_to_write:
            #     for col_idx, cell_value in enumerate(row_data, 1):
            #         col_letter = get_column_letter(col_idx)
            #         ws[f"{col_letter}{startrow}"] = cell_value
            #     startrow += 1

            # # Write the data to the specified sheet_name
            for row_data in data_to_write:
                ws.append(row_data)
            # Save the changes
            wb.save(filename)

            print(f'DataFrame successfully written to sheet "{sheet_name}" in Excel file "{filename}".')
        except Exception as e:
            print(f'Error: {e}')

    else:
        print(f'Excel file "{filename}" does not exist. Creating new file with sheet "{sheet_name}".')
        dataframe.to_excel(filename, sheet_name,index=False)
        
    print(sheet_name,' saved!')
    return