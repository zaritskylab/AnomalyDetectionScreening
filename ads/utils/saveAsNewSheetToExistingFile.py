import pandas as pd
import openpyxl as pxl
from openpyxl.utils import get_column_letter
import os

# ------------------------------------------------------

# Save the input dataframe to the specified sheet name of filename file
def saveAsNewSheetToExistingFile(filename,newDF,newSheetName):
    if os.path.exists(filename):

        excel_book = pxl.load_workbook(filename)
        
#         print(excel_book.sheetnames)
#         if 'cp-cd' in excel_book.sheetnames:
#             print('ghalate')

        if newSheetName in excel_book.sheetnames:
            del excel_book[newSheetName]
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Your loaded workbook is set as the "base of work"
            writer.book = excel_book

            # Loop through the existing worksheets in the workbook and map each title to\
            # the corresponding worksheet (that is, a dictionary where the keys are the\
            # existing worksheets' names and the values are the actual worksheets)
#             print(excel_book.worksheets)
            writer.sheets = {worksheet.title: worksheet for worksheet in excel_book.worksheets if newSheetName not in worksheet}

            # Write the new data to the file without overwriting what already exists
            newDF.to_excel(writer, newSheetName)

            # Save the file
            writer.save()
    else:
        newDF.to_excel(filename, newSheetName)
        
    print(newSheetName,' saved!')
    return


# ------------------------------------------------------

# saveDF_to_CSV_GZ_no_timestamp
def saveDF_to_CSV_GZ_no_timestamp(df,filename):
    from gzip import GzipFile
    from io import TextIOWrapper
    with TextIOWrapper(GzipFile(filename, 'w', mtime=0), encoding='utf-8') as fd:
        df.to_csv(fd,index=False,compression='gzip')
        
    return


# Save the input dataframe to the specified sheet name of filename file
def saveAsNewSheetToExistingFile(filename,newDF,newSheetName):
    if os.path.exists(filename):

        excel_book = pxl.load_workbook(filename)
        
#         print(excel_book.sheetnames)
#         if 'cp-cd' in excel_book.sheetnames:
#             print('ghalate')

        if newSheetName in excel_book.sheetnames:
            del excel_book[newSheetName]
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Your loaded workbook is set as the "base of work"
            writer.book = excel_book

            # Loop through the existing worksheets in the workbook and map each title to\
            # the corresponding worksheet (that is, a dictionary where the keys are the\
            # existing worksheets' names and the values are the actual worksheets)
#             print(excel_book.worksheets)
            writer.sheets = {worksheet.title: worksheet for worksheet in excel_book.worksheets if newSheetName not in worksheet}

            # Write the new data to the file without overwriting what already exists
            newDF.to_excel(writer, newSheetName)

            # Save the file
            writer.save()
    else:
        newDF.to_excel(filename, newSheetName)
        
    print(newSheetName,' saved!')
    return

def write_dataframe_to_excel_old(filename, dataframe, sheet_name):

    if os.path.exists(filename):
        # Load the existing Excel file
        excel_file = pd.ExcelFile(filename)
        # writer = pd.ExcelWriter(filename)
        writer = pd.ExcelWriter(filename, engine='openpyxl')

        writer.book = excel_file.book

        # Check if the sheet_name already exists
        if sheet_name in excel_file.sheet_names:
            print(f'Sheet "{sheet_name}" already exists in the Excel file. Adding new data to the existing sheet.')
            startrow = excel_file.parse(sheet_name).shape[0] + 1
        else:
            startrow = 0

        # Write the DataFrame to the specified sheet_name
        dataframe.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False)

        # Save the changes and close the writer
        writer.save()
        writer.close()


        print(f'DataFrame successfully written to sheet "{sheet_name}" in Excel file "{filename}".')
    else:
        # If the file does not exist, create a new Excel file with the given sheet
        dataframe.to_excel(filename, sheet_name=sheet_name, index=False)
        print(f'Excel file "{filename}" created with sheet "{sheet_name}".')

import pandas as pd


def write_dataframe_to_excel_old(filename, sheet_name, dataframe):
    try:
        # Load the existing Excel file (if it exists)
        excel_file = pd.ExcelFile(filename)

        # Create a Pandas ExcelWriter object with openpyxl engine
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            writer.book = excel_file.book if excel_file else None

            # Check if the sheet_name already exists
            if sheet_name in writer.book.sheetnames:
                print(f'Sheet "{sheet_name}" already exists in the Excel file. Adding new data to the existing sheet.')
                startrow = writer.book[sheet_name].max_row + 1
            else:
                startrow = 0

            # Write the DataFrame to the specified sheet_name
            dataframe.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False)

            # Save the changes
            writer.save()

        print(f'DataFrame successfully written to sheet "{sheet_name}" in Excel file "{filename}".')
    except FileNotFoundError:
        # If the file does not exist, create a new Excel file with the given sheet
        dataframe.to_excel(filename, sheet_name=sheet_name, index=False)
        print(f'Excel file "{filename}" created with sheet "{sheet_name}".')

    

def write_dataframe_to_excel(filename, sheet_name, dataframe):

    if os.path.exists(filename):
        print(f'file {filename} exists')
        try:
            # Load the existing Excel file or create a new one
            try:
                wb = pxl.load_workbook(filename)
                sheet_exists = sheet_name in wb.sheetnames
            except FileNotFoundError:
                wb = pxl.load_workbook(filename)
                sheet_exists = False

            # Check if the sheet_name already exists
            if sheet_exists:
                print(f'Sheet "{sheet_name}" already exists in the Excel file. Adding new data to the existing sheet.')
                ws = wb[sheet_name]
                startrow = ws.max_row + 1
            else:
                ws = wb.create_sheet(sheet_name)
                startrow = 1

            # Convert DataFrame to a list of lists to write to Excel
            data_to_write = dataframe.values.tolist()

                                # Write the column headers to the sheet
            if not sheet_exists:
                for col_idx, column in enumerate(dataframe.columns, 1):
                    col_letter = get_column_letter(col_idx)
                    ws[f"{col_letter}{startrow}"] = column
                startrow += 1

                    # Write the data to the specified sheet_name
            for row_data in data_to_write:
                for col_idx, cell_value in enumerate(row_data, 1):
                    col_letter = get_column_letter(col_idx)
                    ws[f"{col_letter}{startrow}"] = cell_value
                startrow += 1

            # # Write the data to the specified sheet_name
            for row_data in data_to_write:
                ws.append(row_data)
            # Save the changes
            wb.save(filename)

            print(f'DataFrame successfully written to sheet "{sheet_name}" in Excel file "{filename}".')
        except Exception as e:
            print(f'Error: {e}')

    else:
        print(f'Excel file "{filename}" does not exist. Creating new file with sheet "{sheet_name}".')
        dataframe.to_excel(filename, sheet_name)
        
    print(sheet_name,' saved!')
    return